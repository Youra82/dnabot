#!/usr/bin/env python3
# analysis/interactive_chart_momentum_exit.py
# Interaktiver Candlestick-Chart mit momentum_exit-Trade-Signalen
#
# Zeigt:
#   - OHLCV-Candlesticks
#   - Entry-Marker (▲ LONG grün / ▼ SHORT orange)
#   - Exit-Marker  (● WIN cyan / ✗ LOSS rot / ■ TIMEOUT grau)
#   - SL- und Trailing-Aktivierungs-Linien pro Trade
#   - Equity-Kurve (rechte Y-Achse)
#   - Volumen-Panel
#
# Datenquelle: gespeicherte backtest_*_momentum_exit.json-Ergebnisse (siehe
# backtest_momentum_exit.py) fuer die Trades + frisch geladene OHLCV-Daten
# fuer den Candlestick-Hintergrund. Wiederbelebung von interactive_chart.py
# (Genome-System, beim Cleanup 2026-08-24 entfernt) -- deutlich vereinfacht:
# keine Regime-Hintergrundfarben, kein ATR/ADX/Score/Body-ATR-Panel (alles
# genome-spezifisch, hat bei momentum_exit keine Entsprechung -- der Bot hat
# bewusst KEINEN Regime-Filter, siehe momentum_exit_logic.py-Docstring).
#
# Output: HTML-Datei in artifacts/tmp/ (öffnet im Browser).

import os
import sys
import json
from datetime import datetime, timezone

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')

import pandas as pd

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))

RESULTS_DIR = os.path.join(PROJECT_ROOT, 'artifacts', 'results')


def _clean_timeframe(tf: str) -> str:
    return tf[:-len('_momentum_exit')] if tf.endswith('_momentum_exit') else tf


def select_pairs() -> list:
    """Zeigt alle Pairs aus den momentum_exit-Backtest-Ergebnissen mit PnL%."""
    pairs = []
    if os.path.isdir(RESULTS_DIR):
        for fname in sorted(os.listdir(RESULTS_DIR)):
            if not fname.startswith('backtest_') or not fname.endswith('_momentum_exit.json'):
                continue
            try:
                with open(os.path.join(RESULTS_DIR, fname), encoding='utf-8') as f:
                    d = json.load(f)
                pairs.append((d['market'], _clean_timeframe(d['timeframe']),
                              d.get('stats', {}).get('total_pnl_pct')))
            except Exception:
                continue

    if not pairs:
        print("Keine momentum_exit-Backtest-Ergebnisse gefunden. Zuerst Mode 1 ausführen.")
        return []

    pairs.sort(key=lambda x: (x[0], x[1]))

    w = 70
    print("\n" + "=" * w)
    print("  Verfügbare Pairs (momentum_exit):  (PnL = gespeicherter Backtest, voller Zeitraum)")
    print("=" * w)
    for i, (sym, tf, pnl) in enumerate(pairs, 1):
        pnl_str = f"  [+{pnl:.1f}%]" if pnl and pnl > 0 else (f"  [{pnl:.1f}%]" if pnl is not None else "")
        safe = sym.replace('/', '').replace(':', '')
        print(f"  {i:2d}) {safe}_{tf}{pnl_str}")
    print("=" * w)

    print("\n  Wähle Pair(s):")
    print("  Einzeln: z.B. '1' oder '5'")
    print("  Mehrfach: z.B. '1,3,5' oder '1 3 5'")
    raw = input("\n  Auswahl: ").strip()

    selected = []
    for token in raw.replace(',', ' ').split():
        try:
            idx = int(token)
            if 1 <= idx <= len(pairs):
                sym_tf = (pairs[idx - 1][0], pairs[idx - 1][1])
                if sym_tf not in selected:
                    selected.append(sym_tf)
        except ValueError:
            pass

    if not selected:
        print("Ungültige Auswahl.")
    return selected


def _load_saved_trades(symbol: str, timeframe: str) -> tuple:
    """Laedt Trades + Stats aus dem gespeicherten backtest_*_momentum_exit.json."""
    safe = f"{symbol.replace('/', '').replace(':', '')}_{timeframe}_momentum_exit"
    path = os.path.join(RESULTS_DIR, f"backtest_{safe}.json")
    if not os.path.exists(path):
        return [], {}
    with open(path, encoding='utf-8') as f:
        data = json.load(f)
    return data.get('trades', []), data.get('stats', {})


def create_chart(symbol: str, timeframe: str, df: pd.DataFrame, trades: list,
                  stats: dict, start_capital: float, risk_pct: float = 1.0):
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        print("plotly nicht installiert. Bitte: pip install plotly")
        return None

    fig = make_subplots(
        rows=2, cols=1,
        shared_xaxes=True,
        specs=[[{'secondary_y': True}], [{'secondary_y': False}]],
        vertical_spacing=0.03,
        row_heights=[0.78, 0.22],
        subplot_titles=['', 'Volumen'],
    )

    fig.add_trace(go.Candlestick(
        x=df.index,
        open=df['open'], high=df['high'],
        low=df['low'],   close=df['close'],
        name='OHLC',
        increasing_line_color='#26a69a',
        decreasing_line_color='#ef5350',
        showlegend=True,
    ), row=1, col=1, secondary_y=False)

    entry_long_x,  entry_long_y,  entry_long_txt  = [], [], []
    entry_short_x, entry_short_y, entry_short_txt = [], [], []
    exit_win_x,    exit_win_y    = [], []
    exit_loss_x,   exit_loss_y   = [], []
    exit_to_x,     exit_to_y     = [], []

    for t in trades:
        et = pd.to_datetime(t['entry_time'])
        xt = pd.to_datetime(t['exit_time'])
        tip = (
            f"seq_len: {t.get('seq_len', '?')}<br>"
            f"SL: {t['sl_price']:.4f} | Trail-Akt.: {t['tp_price']:.4f}<br>"
            f"SL-Distanz: {t.get('sl_pct', 0):.2f}% | Bewegung: {t.get('pnl_pct', 0):+.2f}%"
        )

        if t['direction'] == 'LONG':
            entry_long_x.append(et); entry_long_y.append(t['entry_price']); entry_long_txt.append(tip)
        else:
            entry_short_x.append(et); entry_short_y.append(t['entry_price']); entry_short_txt.append(tip)

        if t['outcome'] == 'WIN':
            exit_win_x.append(xt);  exit_win_y.append(t['exit_price'])
        elif t['outcome'] == 'LOSS':
            exit_loss_x.append(xt); exit_loss_y.append(t['exit_price'])
        else:
            exit_to_x.append(xt);   exit_to_y.append(t['exit_price'])

        fig.add_shape(
            type='line', x0=et, x1=xt,
            y0=t['sl_price'], y1=t['sl_price'],
            line=dict(color='rgba(239,68,68,0.45)', width=1, dash='dot'),
        )
        fig.add_shape(
            type='line', x0=et, x1=xt,
            y0=t['tp_price'], y1=t['tp_price'],
            line=dict(color='rgba(34,197,94,0.45)', width=1, dash='dot'),
        )

    if entry_long_x:
        fig.add_trace(go.Scatter(
            x=entry_long_x, y=entry_long_y, mode='markers',
            marker=dict(color='#26a69a', symbol='triangle-up', size=14,
                        line=dict(width=1, color='#ffffff')),
            name='Entry Long', text=entry_long_txt,
            hovertemplate='%{text}<extra>Entry Long</extra>',
        ), row=1, col=1, secondary_y=False)

    if entry_short_x:
        fig.add_trace(go.Scatter(
            x=entry_short_x, y=entry_short_y, mode='markers',
            marker=dict(color='#ffa726', symbol='triangle-down', size=14,
                        line=dict(width=1, color='#ffffff')),
            name='Entry Short', text=entry_short_txt,
            hovertemplate='%{text}<extra>Entry Short</extra>',
        ), row=1, col=1, secondary_y=False)

    if exit_win_x:
        fig.add_trace(go.Scatter(
            x=exit_win_x, y=exit_win_y, mode='markers',
            marker=dict(color='#00bcd4', symbol='circle', size=11,
                        line=dict(width=1, color='#ffffff')),
            name='Exit Trailing/TP ✓',
        ), row=1, col=1, secondary_y=False)

    if exit_loss_x:
        fig.add_trace(go.Scatter(
            x=exit_loss_x, y=exit_loss_y, mode='markers',
            marker=dict(color='#ef5350', symbol='x', size=11,
                        line=dict(width=2, color='#ef5350')),
            name='Exit SL ✗',
        ), row=1, col=1, secondary_y=False)

    if exit_to_x:
        fig.add_trace(go.Scatter(
            x=exit_to_x, y=exit_to_y, mode='markers',
            marker=dict(color='#9e9e9e', symbol='square', size=9),
            name='Exit Timeout',
        ), row=1, col=1, secondary_y=False)

    # Equity-Kurve (rechte Y-Achse)
    sorted_trades = sorted(trades, key=lambda t: str(t.get('entry_time', '')))
    equity = start_capital
    eq_times = [df.index[0]]
    eq_vals  = [start_capital]

    for t in sorted_trades:
        risk_amount = equity * (risk_pct / 100.0)
        outcome     = t.get('outcome', 'LOSS')
        if outcome == 'LOSS':
            equity -= risk_amount
        else:
            sl_pct_t = max(t.get('sl_pct', 1.0), 0.01)
            equity  += risk_amount * (t.get('pnl_pct', 0.0) / sl_pct_t)
        eq_times.append(pd.to_datetime(t['entry_time']))
        eq_vals.append(equity)

    if len(eq_vals) > 1:
        fig.add_trace(go.Scatter(
            x=eq_times, y=eq_vals,
            name='Equity',
            line=dict(color='#5c9bd6', width=1.5),
            hovertemplate='Equity: %{y:.2f} USDT<extra></extra>',
        ), row=1, col=1, secondary_y=True)

    if 'volume' in df.columns:
        vol_colors = ['#26a69a' if c >= o else '#ef5350'
                      for c, o in zip(df['close'], df['open'])]
        fig.add_trace(go.Bar(
            x=df.index, y=df['volume'],
            marker_color=vol_colors,
            name='Volumen', showlegend=False, opacity=0.65,
            hovertemplate='Vol: %{y:,.0f}<extra></extra>',
        ), row=2, col=1)

    title = (
        f"DNABOT momentum_exit | {symbol} ({timeframe}) | "
        f"Trades: {stats.get('total_trades', len(trades))} | "
        f"WR: {stats.get('win_rate', 0):.1%} | "
        f"PnL: {stats.get('total_pnl_pct', 0):+.1f}%"
    )
    fig.update_layout(
        title=dict(text=title, font=dict(size=13), x=0.5, xanchor='center'),
        height=850, hovermode='x unified', template='plotly_dark',
        dragmode='zoom',
        xaxis2=dict(rangeslider=dict(visible=True), fixedrange=False),
        legend=dict(orientation='h', yanchor='bottom', y=1.03, xanchor='center', x=0.5),
    )
    fig.update_xaxes(rangeslider_visible=False, row=1, col=1)
    fig.update_yaxes(title_text='Preis',  row=1, col=1, secondary_y=False)
    fig.update_yaxes(title_text='Equity (USDT)', row=1, col=1, secondary_y=True)
    fig.update_yaxes(title_text='Vol',    row=2, col=1)

    return fig


def generate_trades_excel(symbol: str, timeframe: str, trades: list,
                           start_capital: float, risk_pct: float, leverage: int = 1):
    """Excel-Trade-Log fuer EIN Pair (dasselbe Format wie run_manual_portfolio_
    momentum_exit.py/run_portfolio_optimizer_momentum_exit.py, hier nur auf
    ein einzelnes Pair statt ein Portfolio bezogen -- passend zum Einzel-Chart)."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl nicht installiert — Excel übersprungen. (pip install openpyxl)")
        return None

    sorted_trades = sorted(trades, key=lambda t: str(t.get('entry_time', '')))
    equity = start_capital
    rows   = []
    for i, t in enumerate(sorted_trades):
        equity_before = equity
        sl_pct        = max(t.get('sl_pct', 1.0), 0.01)
        risk_amount   = min(equity_before * (risk_pct / 100.0), 200_000.0 * (sl_pct / 100.0))
        outcome       = t.get('outcome', 'LOSS')
        if outcome == 'LOSS':
            pnl = -risk_amount
        else:
            pnl = risk_amount * (t.get('pnl_pct', 0.0) / sl_pct)
        equity += pnl

        raw_position = risk_amount / max(sl_pct / 100.0, 0.0001)
        max_position = min(equity_before * max(leverage, 1), 200_000.0)
        margin       = min(raw_position, max_position) / max(leverage, 1)

        ergebnis = 'Trailing/TP erreicht' if outcome == 'WIN' else ('SL erreicht' if outcome == 'LOSS' else 'Timeout')
        rows.append({
            'Nr':                 i + 1,
            'Datum':              str(t.get('entry_time', ''))[:16].replace('T', ' '),
            'Coin':               symbol.split('/')[0],
            'Timeframe':          timeframe,
            'Richtung':           t.get('direction', '?'),
            'Ergebnis':           ergebnis,
            'Reale Bewegung (%)': round(t.get('pnl_pct', 0.0), 4),
            'Riskiert (USDT)':    round(risk_amount, 4),
            'Marge (USDT)':       round(margin, 4),
            'PnL (USDT)':         round(pnl, 4),
            'Gesamtkapital':      round(equity, 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'

    header_fill  = PatternFill('solid', fgColor='1E3A5F')
    win_fill     = PatternFill('solid', fgColor='D6F4DC')
    loss_fill    = PatternFill('solid', fgColor='FAD7D7')
    timeout_fill = PatternFill('solid', fgColor='FFF3CC')
    alt_fill     = PatternFill('solid', fgColor='F2F2F2')
    thin_border  = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
    )

    headers = list(rows[0].keys()) if rows else []
    col_widths = {
        'Nr': 6, 'Datum': 18, 'Coin': 10, 'Timeframe': 12,
        'Richtung': 10, 'Ergebnis': 18, 'Reale Bewegung (%)': 20,
        'Riskiert (USDT)': 16, 'Marge (USDT)': 14, 'PnL (USDT)': 14, 'Gesamtkapital': 16,
    }
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 14)
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(rows, 2):
        outcome_val = row['Ergebnis']
        if outcome_val == 'Trailing/TP erreicht':
            fill = win_fill
        elif outcome_val == 'SL erreicht':
            fill = loss_fill
        else:
            fill = timeout_fill if r_idx % 2 == 0 else alt_fill
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=col, value=row[key])
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Reale Bewegung (%)', 'Riskiert (USDT)', 'Marge (USDT)', 'PnL (USDT)', 'Gesamtkapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[r_idx].height = 18

    n = len(rows)
    wins = sum(1 for t in trades if t.get('outcome') == 'WIN')
    summary_row = n + 3
    ws.cell(row=summary_row, column=1, value='Zusammenfassung').font = Font(bold=True, size=11)
    for label, value in [
        ('Trades gesamt', n),
        ('Win-Rate', f"{(wins / n * 100.0) if n else 0:.1f}%"),
        ('Final Equity', f"{equity:.2f} USDT"),
        ('Risiko/Trade', f"{risk_pct}%"),
    ]:
        ws.cell(row=summary_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=value)
        summary_row += 1

    tmp_dir = os.path.join(PROJECT_ROOT, 'artifacts', 'tmp')
    os.makedirs(tmp_dir, exist_ok=True)
    safe_name   = f"{symbol.replace('/', '').replace(':', '')}_{timeframe}"
    output_file = os.path.join(tmp_dir, f"dnabot_momentum_exit_{safe_name}_trades.xlsx")
    wb.save(output_file)
    print(f"  ✅ Excel-Tabelle erstellt: {output_file}")
    return output_file


def run_interactive_chart(settings: dict, secrets: dict):
    from dnabot.utils.exchange import Exchange
    from dnabot.utils.config_loader import HISTORY_DAYS_MAP

    print("\n" + "=" * 60)
    print("  INTERAKTIVE CHARTS (momentum_exit)")
    print("=" * 60)

    selected_pairs = select_pairs()
    if not selected_pairs:
        return

    print()
    start_raw = input("Startdatum (JJJJ-MM-TT) [leer=beliebig]: ").strip()
    end_raw   = input("Enddatum   (JJJJ-MM-TT) [leer=heute]: ").strip()

    cap_raw = input("Startkapital in USDT [Standard: 1000]: ").strip()
    start_capital = float(cap_raw) if cap_raw.replace('.', '').isdigit() else 1000.0

    risk_raw = input("Risiko pro Trade in % [Standard: 1.0]: ").strip()
    try:
        chart_risk_pct = float(risk_raw) if risk_raw else 1.0
    except ValueError:
        chart_risk_pct = 1.0

    tg_raw = input("Per Telegram senden? (j/n) [Standard: n]: ").strip().lower()
    send_tg = tg_raw in ('j', 'y', 'yes')

    accounts = secrets.get('dnabot', [])
    if not accounts:
        print("Kein 'dnabot'-Account in secret.json.")
        return
    exchange = Exchange(accounts[0])
    leverage = int(settings.get('risk_settings', {}).get('leverage', 1))

    tmp_dir = os.path.join(PROJECT_ROOT, 'artifacts', 'tmp')
    os.makedirs(tmp_dir, exist_ok=True)
    generated = []

    for symbol, timeframe in selected_pairs:
        print(f"\n--- {symbol} ({timeframe}) ---")

        trades, stats = _load_saved_trades(symbol, timeframe)
        if not trades:
            print("  Keine gespeicherten Trades gefunden — übersprungen.")
            continue

        history_days = HISTORY_DAYS_MAP.get(timeframe, 730)
        print(f"  Lade {history_days} Tage OHLCV...")
        fetch_end   = datetime.now(timezone.utc)
        fetch_start = fetch_end - pd.Timedelta(days=history_days)
        df = exchange.fetch_historical_ohlcv(
            symbol, timeframe,
            fetch_start.strftime('%Y-%m-%d'),
            fetch_end.strftime('%Y-%m-%d'),
        )
        if df is None or df.empty:
            print("  Keine Daten — übersprungen.")
            continue
        print(f"  {len(df)} Kerzen geladen, {len(trades)} gespeicherte Trades.")

        df_chart     = df.copy()
        trades_chart = trades
        if start_raw:
            try:
                sd = pd.Timestamp(start_raw, tz='UTC')
                df_chart     = df_chart[df_chart.index >= sd]
                trades_chart = [t for t in trades_chart if str(t.get('entry_time', '')) >= start_raw]
            except Exception:
                pass
        if end_raw:
            try:
                ed = pd.Timestamp(end_raw + ' 23:59:59', tz='UTC')
                df_chart     = df_chart[df_chart.index <= ed]
                trades_chart = [t for t in trades_chart if str(t.get('entry_time', '')) <= end_raw + ' 23:59:59']
            except Exception:
                pass

        if df_chart.empty:
            print("  Kein Datenbereich nach Datumsfilter — übersprungen.")
            continue

        print("  Erstelle Chart...")
        fig = create_chart(symbol, timeframe, df_chart, trades_chart, stats,
                           start_capital, risk_pct=chart_risk_pct)
        if fig is None:
            continue

        safe_name   = f"{symbol.replace('/', '').replace(':', '')}_{timeframe}"
        output_file = os.path.join(tmp_dir, f"dnabot_momentum_exit_{safe_name}.html")
        fig.write_html(output_file)
        print(f"  ✅ Chart gespeichert: {output_file}")

        excel_file = generate_trades_excel(symbol, timeframe, trades_chart,
                                            start_capital, chart_risk_pct, leverage)

        generated.append((symbol, timeframe, output_file, excel_file))

    print(f"\n✅ {len(generated)} Chart(s) generiert!")

    if send_tg and generated:
        tg = secrets.get('telegram', {})
        accounts = secrets.get('dnabot', [])
        acc = accounts[0] if accounts else {}
        bot_token = acc.get('telegram_bot_token', '') or tg.get('bot_token', '')
        chat_id   = acc.get('telegram_chat_id', '')   or tg.get('chat_id', '')
        if bot_token and chat_id:
            try:
                from dnabot.utils.telegram import send_document
                for sym, tf, path, excel_path in generated:
                    send_document(bot_token, chat_id, path, caption=f"dnabot momentum_exit Chart: {sym} {tf}")
                    print(f"  ✅ Telegram: {sym} {tf} Chart gesendet.")
                    if excel_path:
                        send_document(bot_token, chat_id, excel_path,
                                      caption=f"dnabot momentum_exit Trades: {sym} {tf}")
                        print(f"  ✅ Telegram: {sym} {tf} Excel gesendet.")
            except Exception as e:
                print(f"  Telegram-Fehler: {e}")
        else:
            print("  Telegram nicht konfiguriert (bot_token/chat_id fehlt).")


def main():
    settings_path = os.path.join(PROJECT_ROOT, 'settings.json')
    secret_path   = os.path.join(PROJECT_ROOT, 'secret.json')
    with open(settings_path, encoding='utf-8') as f:
        settings = json.load(f)
    with open(secret_path, encoding='utf-8') as f:
        secrets = json.load(f)
    run_interactive_chart(settings, secrets)


if __name__ == '__main__':
    main()
