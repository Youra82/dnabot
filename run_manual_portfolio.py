#!/usr/bin/env python3
# run_manual_portfolio.py
# Manuelle Portfolio-Simulation: Nutzer wählt Pairs, Bot simuliert kombinierten Kapital-Pool.

import os
import sys
import json
import argparse
from datetime import datetime, timezone

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))

RESULTS_DIR   = os.path.join(PROJECT_ROOT, 'artifacts', 'results')
SETTINGS_PATH = os.path.join(PROJECT_ROOT, 'settings.json')

G   = '\033[0;32m'
Y   = '\033[1;33m'
R   = '\033[0;31m'
C   = '\033[0;36m'
B   = '\033[1;37m'
NC  = '\033[0m'

RR_RATIO          = 2.0
MAX_NOTIONAL_USDT = 200_000.0


def load_all_results(start_date=None, end_date=None):
    results = []
    if not os.path.isdir(RESULTS_DIR):
        return results

    sd = datetime.fromisoformat(start_date).replace(tzinfo=timezone.utc) if start_date else None
    ed = datetime.fromisoformat(end_date + 'T23:59:59').replace(tzinfo=timezone.utc) if end_date else None

    for fname in sorted(os.listdir(RESULTS_DIR)):
        if not fname.startswith('backtest_') or not fname.endswith('.json'):
            continue
        path = os.path.join(RESULTS_DIR, fname)
        try:
            with open(path) as f:
                data = json.load(f)
        except Exception:
            continue

        trades = data.get('trades', [])
        if sd or ed:
            filtered = []
            for t in trades:
                ts = t.get('entry_time', '')
                if not ts:
                    continue
                try:
                    t_dt = datetime.fromisoformat(str(ts))
                    if t_dt.tzinfo is None:
                        t_dt = t_dt.replace(tzinfo=timezone.utc)
                except Exception:
                    continue
                if sd and t_dt < sd:
                    continue
                if ed and t_dt > ed:
                    continue
                filtered.append(t)
            trades = filtered

        results.append({
            'market':    data['market'],
            'timeframe': data['timeframe'],
            'trades':    trades,
        })

    return results


def simulate_portfolio(pair_results, capital, risk_pct):
    if not pair_results:
        return {'total_pnl_pct': 0.0, 'final_equity': capital,
                'max_dd': 0.0, 'n_trades': 0, 'win_rate': 0.0}

    all_trades = []
    for pr in pair_results:
        for t in pr['trades']:
            all_trades.append({
                'market':     pr['market'],
                'timeframe':  pr['timeframe'],
                'outcome':    t.get('outcome', 'LOSS'),
                'pnl_pct':    t.get('pnl_pct', 0.0),
                'sl_pct':     t.get('sl_pct', 1.0),
                'entry_time': str(t.get('entry_time', '')),
            })

    all_trades.sort(key=lambda t: t['entry_time'])

    equity = capital
    peak   = equity
    max_dd = 0.0
    wins   = 0

    for t in all_trades:
        sl_pct      = max(t['sl_pct'], 0.01)
        risk_amount = min(equity * (risk_pct / 100.0), MAX_NOTIONAL_USDT * (sl_pct / 100.0))
        outcome     = t['outcome']

        if outcome == 'WIN':
            pnl = risk_amount * RR_RATIO
            wins += 1
        elif outcome == 'LOSS':
            pnl = -risk_amount
        else:
            pnl = risk_amount * (t['pnl_pct'] / sl_pct)

        equity += pnl
        if equity > peak:
            peak = equity
        if peak > 0:
            dd = (peak - equity) / peak * 100.0
            if dd > max_dd:
                max_dd = dd

    n = len(all_trades)
    total_pnl_pct = (equity - capital) / capital * 100.0 if capital > 0 else 0.0

    return {
        'total_pnl_pct': total_pnl_pct,
        'final_equity':  equity,
        'max_dd':        max_dd,
        'n_trades':      n,
        'win_rate':      wins / n if n > 0 else 0.0,
    }


def compute_single_stats(trades, capital, risk_pct):
    return simulate_portfolio(
        [{'market': '', 'timeframe': '', 'trades': trades}],
        capital, risk_pct
    )


def select_pairs(all_results, capital, risk_pct):
    """Zeigt alle Pairs mit PnL% und lässt Nutzer auswählen."""
    # Einzel-Stats berechnen
    pairs_with_stats = []
    for r in all_results:
        st = compute_single_stats(r['trades'], capital, risk_pct)
        pairs_with_stats.append({**r, 'stats': st})

    # Nach PnL% sortieren
    pairs_with_stats.sort(key=lambda x: x['stats']['total_pnl_pct'], reverse=True)

    w = 72
    print(f"\n{'=' * w}")
    print(f"{B}  Verfügbare Pairs{NC}")
    print(f"  {'Nr':<4} {'Markt':<24} {'TF':<6} {'Trades':>7} {'WR':>7} {'PnL%':>9} {'MaxDD':>8}")
    print(f"  {'-' * (w - 2)}")

    for i, pr in enumerate(pairs_with_stats, 1):
        st = pr['stats']
        if st['n_trades'] == 0:
            continue
        pnl_col = G if st['total_pnl_pct'] > 0 else R
        wr_col  = G if st['win_rate'] >= 0.50 else (Y if st['win_rate'] >= 0.43 else R)
        sign    = '+' if st['total_pnl_pct'] >= 0 else ''
        print(
            f"  {i:<4} {pr['market']:<24} {pr['timeframe']:<6} {st['n_trades']:>7} "
            f"{wr_col}{st['win_rate']:>6.1%}{NC} "
            f"{pnl_col}{sign}{st['total_pnl_pct']:>7.1f}%{NC} "
            f"{st['max_dd']:>7.1f}%"
        )

    print(f"{'=' * w}")
    print(f"\n  Eingabe: Nummern kommagetrennt (z.B. {Y}1,3,5{NC}) oder {Y}alle{NC}")

    try:
        sel = input("  Auswahl: ").strip()
    except (EOFError, KeyboardInterrupt):
        print()
        return []

    if not sel or sel.lower() in ('alle', 'all', 'a'):
        return [p for p in pairs_with_stats if p['stats']['n_trades'] > 0]

    selected = []
    valid_pairs = [p for p in pairs_with_stats if p['stats']['n_trades'] > 0]
    try:
        indices = [int(x.strip()) - 1 for x in sel.split(',')]
        for idx in indices:
            if 0 <= idx < len(valid_pairs):
                selected.append(valid_pairs[idx])
    except ValueError:
        print(f"  {R}Ungültige Eingabe.{NC}")
        return []

    return selected


def build_telegram_report(selected, pm, capital, risk_pct, start_date, end_date):
    date_range = f"{start_date or '...'} → {end_date or 'heute'}"
    sign = '+' if pm['total_pnl_pct'] >= 0 else ''
    lines = [
        "dnabot — Manuelle Portfolio-Simulation",
        f"Zeitraum: {date_range}",
        f"Kapital: {capital:.0f} USDT | Risiko: {risk_pct}%",
        "",
        f"Ausgewählte Pairs ({len(selected)}):",
    ]
    for pr in selected:
        st = pr['stats']
        s = '+' if st['total_pnl_pct'] >= 0 else ''
        lines.append(
            f"  {pr['market']} {pr['timeframe']} | "
            f"{st['n_trades']} Trades | WR {st['win_rate']:.1%} | PnL {s}{st['total_pnl_pct']:.1f}%"
        )
    lines += [
        "",
        f"Portfolio gesamt:",
        f"  Trades:       {pm['n_trades']}",
        f"  Win-Rate:     {pm['win_rate']:.1%}",
        f"  PnL:          {sign}{pm['total_pnl_pct']:.1f}%",
        f"  Final Equity: {pm['final_equity']:.2f} USDT",
        f"  Max Drawdown: {pm['max_dd']:.1f}%",
    ]
    return "\n".join(lines)


def print_result(selected, pm, capital, risk_pct, start_date, end_date):
    w = 72
    date_range = f" | {start_date or '...'} → {end_date or 'heute'}"
    print(f"\n{'=' * w}")
    print(f"{B}  dnabot — Manuelle Portfolio-Simulation{NC}")
    print(f"  Zeitraum:{date_range}")
    print(f"  Kapital: {capital:.0f} USDT | Risiko/Trade: {risk_pct}% (gemeinsamer Pool)")
    print(f"{'=' * w}")

    print(f"\n  {G}Ausgewählte Pairs — {len(selected)} Pair(s){NC}")
    print(f"  {'Markt':<24} {'TF':<6} {'Trades':>7} {'WR':>7} {'PnL%':>9} {'MaxDD':>8}")
    print(f"  {'-' * (w - 2)}")

    for pr in selected:
        st      = pr['stats']
        pnl_col = G if st['total_pnl_pct'] > 0 else R
        wr_col  = G if st['win_rate'] >= 0.50 else (Y if st['win_rate'] >= 0.43 else R)
        sign    = '+' if st['total_pnl_pct'] >= 0 else ''
        print(
            f"  {pr['market']:<24} {pr['timeframe']:<6} {st['n_trades']:>7} "
            f"{wr_col}{st['win_rate']:>6.1%}{NC} "
            f"{pnl_col}{sign}{st['total_pnl_pct']:>7.1f}%{NC} "
            f"{st['max_dd']:>7.1f}%"
        )

    pnl_col = G if pm['total_pnl_pct'] > 0 else R
    sign    = '+' if pm['total_pnl_pct'] >= 0 else ''
    print(f"\n  {'─' * (w - 2)}")
    print(f"  {B}Portfolio gesamt (gemeinsamer Kapital-Pool, alle Trades kompoundiert):{NC}")
    print(f"  Trades total:  {pm['n_trades']}")
    print(f"  Win-Rate:      {pm['win_rate']:.1%}")
    print(f"  PnL:           {pnl_col}{sign}{pm['total_pnl_pct']:.1f}%{NC}")
    print(f"  Final Equity:  {pm['final_equity']:.2f} USDT")
    print(f"  Max Drawdown:  {pm['max_dd']:.1f}%")
    print(f"{'=' * w}\n")


def _get_telegram_credentials():
    secret_path = os.path.join(PROJECT_ROOT, 'secret.json')
    try:
        with open(secret_path) as f:
            secrets = json.load(f)
        accounts  = secrets.get('dnabot', [])
        acc       = accounts[0] if accounts else {}
        bot_token = acc.get('telegram_bot_token', '') or secrets.get('telegram', {}).get('bot_token', '')
        chat_id   = acc.get('telegram_chat_id', '')   or secrets.get('telegram', {}).get('chat_id', '')
        if bot_token and chat_id:
            return bot_token, chat_id
    except Exception:
        pass
    # Fallback: settings.json (legacy)
    try:
        with open(SETTINGS_PATH) as f:
            settings = json.load(f)
        cfg = settings.get('dnabot', settings)
        bt  = cfg.get('telegram_bot_token') or settings.get('telegram_bot_token', '')
        ci  = cfg.get('telegram_chat_id')   or settings.get('telegram_chat_id', '')
        if bt and ci:
            return bt, ci
    except Exception:
        pass
    return None, None


def send_telegram(report_text):
    bot_token, chat_id = _get_telegram_credentials()
    if not bot_token or not chat_id:
        print(f"  {Y}Kein Telegram-Token/Chat-ID in secret.json — übersprungen.{NC}")
        return
    sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))
    from dnabot.utils.telegram import send_message
    send_message(bot_token, chat_id, report_text)
    print(f"  {G}✓ Telegram-Nachricht gesendet.{NC}")


def generate_portfolio_equity_chart(selected: list, pm: dict,
                                     start_date: str, end_date: str,
                                     capital: float, risk_pct: float):
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        print(f"{R}  plotly nicht installiert — Chart übersprungen.{NC}")
        return

    bot_token, chat_id = _get_telegram_credentials()

    all_trades = []
    for pr in selected:
        for t in pr['trades']:
            all_trades.append({
                'market':     pr['market'],
                'timeframe':  pr['timeframe'],
                'outcome':    t.get('outcome', 'LOSS'),
                'pnl_pct':    t.get('pnl_pct', 0.0),
                'sl_pct':     t.get('sl_pct', 1.0),
                'entry_time': str(t.get('entry_time', '')),
            })
    all_trades.sort(key=lambda t: t['entry_time'])

    if not all_trades:
        print(f"  {Y}Keine Trades vorhanden — Chart übersprungen.{NC}")
        return

    equity   = capital
    peak     = equity
    eq_times = [all_trades[0]['entry_time']]
    eq_vals  = [equity]
    wins     = 0

    for t in all_trades:
        sl_pct      = max(t['sl_pct'], 0.01)
        risk_amount = min(equity * (risk_pct / 100.0), MAX_NOTIONAL_USDT * (sl_pct / 100.0))
        if t['outcome'] == 'WIN':
            equity += risk_amount * RR_RATIO
            wins   += 1
        elif t['outcome'] == 'LOSS':
            equity -= risk_amount
        else:
            equity += risk_amount * (t['pnl_pct'] / sl_pct)
        if equity > peak:
            peak = equity
        eq_times.append(t['entry_time'])
        eq_vals.append(round(equity, 2))

    n       = len(all_trades)
    wr      = wins / n if n > 0 else 0.0
    pnl_pct = (equity - capital) / capital * 100.0 if capital > 0 else 0.0
    max_dd  = pm['max_dd']
    sign    = '+' if pnl_pct >= 0 else ''

    PAIR_COLORS = [
        '#f59e0b', '#8b5cf6', '#ec4899', '#14b8a6',
        '#f97316', '#84cc16', '#06b6d4', '#a78bfa',
    ]
    pair_equity_traces = []
    for idx, pr in enumerate(selected):
        pair_trades = sorted(pr['trades'], key=lambda t: str(t.get('entry_time', '')))
        peq    = capital
        ptimes = [str(pair_trades[0].get('entry_time', ''))] if pair_trades else []
        pvals  = [peq]
        for t in pair_trades:
            slp = max(t.get('sl_pct', 1.0), 0.01)
            ra  = min(peq * (risk_pct / 100.0), MAX_NOTIONAL_USDT * (slp / 100.0))
            out = t.get('outcome', 'LOSS')
            if out == 'WIN':
                peq += ra * RR_RATIO
            elif out == 'LOSS':
                peq -= ra
            else:
                peq += ra * (t.get('pnl_pct', 0.0) / slp)
            ptimes.append(str(t.get('entry_time', '')))
            pvals.append(round(peq, 2))
        label = f"{pr['market'].split('/')[0]}/{pr['timeframe']}"
        pair_equity_traces.append((ptimes, pvals, label, PAIR_COLORS[idx % len(PAIR_COLORS)]))

    date_range = f"{start_date or '...'} → {end_date or 'heute'}"
    pairs_str  = ', '.join(f"{p['market'].split('/')[0]}/{p['timeframe']}" for p in selected)
    title = (
        f"dnabot Portfolio — {len(selected)} Coins ({pairs_str}) | "
        f"Trades: {n} | WR: {wr:.1%} | "
        f"PnL: {sign}{pnl_pct:.1f}% | "
        f"Final Equity: {equity:.2f} USDT | MaxDD: {max_dd:.1f}%"
    )

    fig = make_subplots(specs=[[{"secondary_y": True}]])

    for ptimes, pvals, label, color in pair_equity_traces:
        fig.add_trace(go.Scatter(
            x=ptimes, y=pvals, mode='lines', name=label,
            line=dict(color=color, width=1), opacity=0.55,
        ), secondary_y=False)

    fig.add_hline(
        y=capital,
        line=dict(color='rgba(100,100,100,0.35)', width=1, dash='dash'),
        annotation_text=f'Start {capital:.0f} USDT',
        annotation_position='top left',
    )

    entry_long_x, entry_long_y, entry_long_txt = [], [], []
    exit_win_x,   exit_win_y                   = [], []
    exit_loss_x,  exit_loss_y                  = [], []
    exit_to_x,    exit_to_y                    = [], []

    for i, t in enumerate(all_trades):
        eq_val = eq_vals[i + 1]
        tip    = f"{t['market']} {t['timeframe']}<br>Equity: {eq_val:.2f} USDT"
        entry_long_x.append(t['entry_time'])
        entry_long_y.append(eq_val)
        entry_long_txt.append(tip)
        if t['outcome'] == 'WIN':
            exit_win_x.append(t['entry_time']);  exit_win_y.append(eq_val)
        elif t['outcome'] == 'LOSS':
            exit_loss_x.append(t['entry_time']); exit_loss_y.append(eq_val)
        else:
            exit_to_x.append(t['entry_time']);   exit_to_y.append(eq_val)

    fig.add_trace(go.Scatter(
        x=eq_times, y=eq_vals, mode='lines',
        name='Portfolio Equity',
        line=dict(color='#2563eb', width=2), opacity=0.75,
    ), secondary_y=True)

    if entry_long_x:
        fig.add_trace(go.Scatter(
            x=entry_long_x, y=entry_long_y, mode='markers',
            marker=dict(color='#16a34a', symbol='triangle-up', size=14,
                        line=dict(width=1, color='#0f5132')),
            name='Entry ▲', text=entry_long_txt,
            hovertemplate='%{text}<extra>Entry</extra>',
        ), secondary_y=True)

    if exit_win_x:
        fig.add_trace(go.Scatter(
            x=exit_win_x, y=exit_win_y, mode='markers',
            marker=dict(color='#22d3ee', symbol='circle', size=11,
                        line=dict(width=1, color='#0e7490')),
            name='Exit TP ✓',
        ), secondary_y=True)

    if exit_loss_x:
        fig.add_trace(go.Scatter(
            x=exit_loss_x, y=exit_loss_y, mode='markers',
            marker=dict(color='#ef4444', symbol='x', size=11,
                        line=dict(width=2, color='#7f1d1d')),
            name='Exit SL ✗',
        ), secondary_y=True)

    if exit_to_x:
        fig.add_trace(go.Scatter(
            x=exit_to_x, y=exit_to_y, mode='markers',
            marker=dict(color='#9ca3af', symbol='square', size=9),
            name='Exit Timeout',
        ), secondary_y=True)

    fig.update_layout(
        title=dict(text=title, font=dict(size=13), x=0.5, xanchor='center'),
        height=750, hovermode='x unified', template='plotly_white',
        dragmode='zoom',
        xaxis=dict(rangeslider=dict(visible=True), fixedrange=False),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
    )
    fig.update_yaxes(title_text='Einzel-Equity (USDT)', secondary_y=False, fixedrange=False)
    fig.update_yaxes(title_text='Portfolio-Equity (USDT)', secondary_y=True,  fixedrange=False)

    output_file = '/tmp/dnabot_portfolio_equity.html'
    fig.write_html(output_file)
    print(f"\n  {G}✓ Portfolio-Chart erstellt: {output_file}{NC}")

    if bot_token and chat_id:
        sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))
        from dnabot.utils.telegram import send_document
        caption = (
            f"dnabot Portfolio-Equity\n"
            f"{date_range} | {len(selected)} Coins | "
            f"PnL: {sign}{pnl_pct:.1f}% | Equity: {equity:.2f} USDT | MaxDD: {max_dd:.1f}%"
        )
        send_document(bot_token, chat_id, output_file, caption=caption)
        print(f"  {G}✓ HTML-Chart via Telegram gesendet.{NC}")
    else:
        print(f"  {Y}Telegram nicht konfiguriert — Chart nur lokal gespeichert.{NC}")


def generate_trades_excel(selected: list, pm: dict, capital: float, risk_pct: float,
                          leverage: int = 1):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"  {Y}openpyxl nicht installiert — Excel übersprungen. (pip install openpyxl){NC}")
        return None

    all_trades = []
    for pr in selected:
        for t in pr['trades']:
            all_trades.append({
                'market':     pr['market'],
                'timeframe':  pr['timeframe'],
                'coin':       pr['market'].split('/')[0],
                'direction':  t.get('direction', '?'),
                'outcome':    t.get('outcome', 'LOSS'),
                'pnl_pct':    t.get('pnl_pct', 0.0),
                'sl_pct':     t.get('sl_pct', 1.0),
                'entry_time': str(t.get('entry_time', '')),
                'exit_time':  str(t.get('exit_time', '')),
            })
    all_trades.sort(key=lambda t: t['entry_time'])

    equity = capital
    rows   = []
    for i, t in enumerate(all_trades):
        equity_before = equity
        sl_pct        = max(t['sl_pct'], 0.01)
        risk_amount   = min(equity_before * (risk_pct / 100.0), MAX_NOTIONAL_USDT * (sl_pct / 100.0))
        if t['outcome'] == 'WIN':
            pnl = risk_amount * RR_RATIO
        elif t['outcome'] == 'LOSS':
            pnl = -risk_amount
        else:
            pnl = risk_amount * (t['pnl_pct'] / sl_pct)
        equity += pnl

        raw_position = risk_amount / max(sl_pct / 100.0, 0.0001)
        max_position = min(equity_before * max(leverage, 1), MAX_NOTIONAL_USDT)
        margin       = min(raw_position, max_position) / max(leverage, 1)

        ergebnis = 'TP erreicht' if t['outcome'] == 'WIN' else ('SL erreicht' if t['outcome'] == 'LOSS' else 'Timeout')
        rows.append({
            'Nr':                    i + 1,
            'Datum':                 t['entry_time'][:16].replace('T', ' '),
            'Coin':                  t['coin'],
            'Timeframe':             t['timeframe'],
            'Richtung':              t['direction'],
            'Ergebnis':              ergebnis,
            'Reale Bewegung (%)':    round(t.get('pnl_pct', 0.0), 4),
            'Riskiert (USDT)':       round(risk_amount, 4),
            'Marge (USDT)':          round(margin, 4),
            'PnL (USDT)':            round(pnl, 4),
            'Gesamtkapital':         round(equity, 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'

    header_fill  = PatternFill('solid', fgColor='1E3A5F')
    win_fill     = PatternFill('solid', fgColor='D6F4DC')
    loss_fill    = PatternFill('solid', fgColor='FAD7D7')
    timeout_fill = PatternFill('solid', fgColor='FFF3CC')
    alt_fill     = PatternFill('solid', fgColor='F2F2F2')
    thin_border  = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    headers = list(rows[0].keys()) if rows else []
    col_widths = {
        'Nr': 6, 'Datum': 18, 'Coin': 10, 'Timeframe': 12,
        'Richtung': 10, 'Ergebnis': 14, 'Reale Bewegung (%)': 20,
        'Riskiert (USDT)': 16, 'Marge (USDT)': 14, 'PnL (USDT)': 14, 'Gesamtkapital': 16,
    }

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 14)
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(rows, 2):
        outcome_val = row['Ergebnis']
        if outcome_val == 'TP erreicht':
            fill = win_fill
        elif outcome_val == 'SL erreicht':
            fill = loss_fill
        else:
            fill = timeout_fill if r_idx % 2 == 0 else alt_fill

        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=col, value=row[key])
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Reale Bewegung (%)', 'Riskiert (USDT)', 'Marge (USDT)', 'PnL (USDT)', 'Gesamtkapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[r_idx].height = 18

    summary_row = len(rows) + 3
    ws.cell(row=summary_row, column=1, value='Zusammenfassung').font = Font(bold=True, size=11)
    for label, value in [
        ('Trades gesamt', pm['n_trades']),
        ('Win-Rate', f"{pm['win_rate']:.1%}"),
        ('PnL', f"+{pm['total_pnl_pct']:.1f}%"),
        ('Final Equity', f"{pm['final_equity']:.2f} USDT"),
        ('Max Drawdown', f"{pm['max_dd']:.1f}%"),
        ('Risiko/Trade', f"{risk_pct}%"),
    ]:
        ws.cell(row=summary_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=value)
        summary_row += 1

    output_file = '/tmp/dnabot_trades.xlsx'
    wb.save(output_file)
    print(f"  {G}✓ Excel-Tabelle erstellt: {output_file}{NC}")
    return output_file


def main():
    parser = argparse.ArgumentParser(description="dnabot Manuelle Portfolio-Simulation")
    parser.add_argument('--capital',    type=float, default=1000.0)
    parser.add_argument('--risk',       type=float, default=1.0)
    parser.add_argument('--start-date', type=str,   default=None)
    parser.add_argument('--end-date',   type=str,   default=None)
    parser.add_argument('--telegram',   action='store_true')
    args = parser.parse_args()

    w = 72
    date_range = f" | {args.start_date or '...'} → {args.end_date or 'heute'}"
    print(f"\n{'─' * w}")
    print(f"{B}  dnabot Manuelle Portfolio-Simulation{NC}")
    print(f"  Kapital: {args.capital:.0f} USDT | Risiko/Trade: {args.risk}%{date_range}")
    print(f"  Modell: Gemeinsamer Kapital-Pool — alle Trades kompoundieren zusammen")
    print(f"{'─' * w}\n")

    print("  Lade Backtest-Ergebnisse ...", end='', flush=True)
    all_results = load_all_results(args.start_date, args.end_date)
    with_trades = [r for r in all_results if len(r['trades']) > 0]

    if not with_trades:
        print(f"\n  {R}Keine Backtest-Ergebnisse gefunden.{NC}")
        print("  Zuerst Mode 1 (Einzel-Backtest) ausführen!\n")
        sys.exit(1)

    print(f" {len(all_results)} Dateien, {len(with_trades)} mit Trades.")

    selected = select_pairs(with_trades, args.capital, args.risk)

    if not selected:
        print(f"\n  {Y}Keine Pairs ausgewählt. Abbruch.{NC}\n")
        sys.exit(0)

    pm = simulate_portfolio(selected, args.capital, args.risk)
    print_result(selected, pm, args.capital, args.risk, args.start_date, args.end_date)

    send_tg = args.telegram
    if not send_tg:
        try:
            ans = input("  Ergebnisse an Telegram senden? (j/n): ").strip().lower()
            send_tg = ans in ('j', 'ja', 'y', 'yes')
        except (EOFError, KeyboardInterrupt):
            pass

    if send_tg:
        report = build_telegram_report(selected, pm, args.capital, args.risk,
                                       args.start_date, args.end_date)
        send_telegram(report)

    # Leverage aus settings.json für Excel (Margin-Berechnung)
    leverage = 1
    try:
        with open(SETTINGS_PATH) as f:
            _s = json.load(f)
        leverage = int(_s.get('risk_settings', {}).get('leverage', 1))
    except Exception:
        pass

    try:
        chart_ans = input("  Interaktive Charts & Excel erstellen und via Telegram senden? (j/n): ").strip().lower()
    except (EOFError, KeyboardInterrupt):
        chart_ans = 'n'
    if chart_ans in ('j', 'ja', 'y', 'yes'):
        generate_portfolio_equity_chart(
            selected, pm, args.start_date, args.end_date, args.capital, args.risk
        )
        excel_file = generate_trades_excel(selected, pm, args.capital, args.risk, leverage)
        if excel_file:
            bot_token, chat_id = _get_telegram_credentials()
            if bot_token and chat_id:
                sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))
                from dnabot.utils.telegram import send_document
                send_document(
                    bot_token, chat_id, excel_file,
                    caption=f"dnabot Trades-Tabelle | {len(selected)} Coins | "
                            f"Risiko: {args.risk}% | {pm['n_trades']} Trades | "
                            f"WR: {pm['win_rate']:.1%} | Final: {pm['final_equity']:.2f} USDT"
                )
                print(f"  {G}✓ Excel via Telegram gesendet.{NC}")


if __name__ == '__main__':
    main()
