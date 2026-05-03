#!/usr/bin/env python3
# run_portfolio_optimizer.py
# Automatische Portfolio-Optimierung (jaegerbot-Style):
#
# Kapital-Modell:
#   - EIN gemeinsamer Kapital-Pool für alle Pairs
#   - Alle Trades laufen chronologisch auf demselben Kapital
#   - Jeder Trade riskiert risk_pct% des AKTUELLEN Equity
#   - Mehr profitable Trades = mehr Kompoundierung = höherer PnL
#   - Nebenläufige Trades werden nacheinander auf das gleiche Equity angewandt
#
# Optimizer:
#   - Exhaustive Suche: testet alle möglichen Pair-Kombinationen
#   - Constraint: max. 1 TF pro Coin (Bitget: 1 Position pro Symbol)
#   - Stoppt wenn keine weitere Verbesserung durch mehr Coins möglich ist

import os
import sys
import json
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.join(PROJECT_ROOT, 'src'))

RESULTS_DIR   = os.path.join(PROJECT_ROOT, 'artifacts', 'results')
SETTINGS_PATH = os.path.join(PROJECT_ROOT, 'settings.json')

G   = '\033[0;32m'
Y   = '\033[1;33m'
R   = '\033[0;31m'
C   = '\033[0;36m'
B   = '\033[1;37m'
NC  = '\033[0m'

RR_RATIO  = 2.0
N_WORKERS = min(os.cpu_count() or 4, 8)


def _get_telegram_credentials():
    secret_path = os.path.join(PROJECT_ROOT, 'secret.json')
    try:
        with open(secret_path) as f:
            secrets = json.load(f)
        accounts  = secrets.get('dnabot', [])
        acc       = accounts[0] if accounts else {}
        bot_token = acc.get('telegram_bot_token', '') or secrets.get('telegram', {}).get('bot_token', '')
        chat_id   = acc.get('telegram_chat_id', '')   or secrets.get('telegram', {}).get('chat_id', '')
        if bot_token and chat_id:
            return bot_token, chat_id
    except Exception:
        pass
    return None, None


def _send_telegram(message: str):
    bot_token, chat_id = _get_telegram_credentials()
    if not bot_token or not chat_id:
        return
    try:
        import requests
        requests.post(
            f"https://api.telegram.org/bot{bot_token}/sendMessage",
            data={'chat_id': chat_id, 'text': message},
            timeout=10,
        )
    except Exception:
        pass


def coin_from_symbol(symbol: str) -> str:
    return symbol.split('/')[0].upper()


def load_all_results(start_date=None, end_date=None):
    results = []
    if not os.path.isdir(RESULTS_DIR):
        return results

    sd = datetime.fromisoformat(start_date).replace(tzinfo=timezone.utc) if start_date else None
    ed = datetime.fromisoformat(end_date + 'T23:59:59').replace(tzinfo=timezone.utc) if end_date else None

    for fname in sorted(os.listdir(RESULTS_DIR)):
        if not fname.startswith('backtest_') or not fname.endswith('.json'):
            continue
        path = os.path.join(RESULTS_DIR, fname)
        try:
            with open(path) as f:
                data = json.load(f)
        except Exception:
            continue

        trades = data.get('trades', [])
        if sd or ed:
            filtered = []
            for t in trades:
                ts = t.get('entry_time', '')
                if not ts:
                    continue
                try:
                    t_dt = datetime.fromisoformat(str(ts))
                    if t_dt.tzinfo is None:
                        t_dt = t_dt.replace(tzinfo=timezone.utc)
                except Exception:
                    continue
                if sd and t_dt < sd:
                    continue
                if ed and t_dt > ed:
                    continue
                filtered.append(t)
            trades = filtered

        results.append({
            'market':    data['market'],
            'timeframe': data['timeframe'],
            'coin':      coin_from_symbol(data['market']),
            'trades':    trades,
            'stats':     data.get('stats', {}),
        })

    return results


def simulate_portfolio(pair_results: list, capital: float, risk_pct: float) -> dict:
    """
    Simuliert ein Portfolio mit GEMEINSAMEM Kapital-Pool.

    Alle Trades aller Pairs werden chronologisch zusammengeführt.
    Jeder Trade riskiert risk_pct% des aktuellen Equity (Kompoundierung).
    Das ermöglicht höhere PnL als Einzel-Pairs durch mehr Trades.
    """
    if not pair_results:
        return {
            'total_pnl_pct': 0.0, 'final_equity': capital,
            'max_dd': 0.0, 'n_trades': 0, 'win_rate': 0.0,
        }

    # Alle Trades zusammenführen und chronologisch sortieren
    all_trades = []
    for pr in pair_results:
        for t in pr['trades']:
            all_trades.append({
                'market':    pr['market'],
                'timeframe': pr['timeframe'],
                'outcome':   t.get('outcome', 'LOSS'),
                'pnl_pct':   t.get('pnl_pct', 0.0),
                'sl_pct':    t.get('sl_pct', 1.0),
                'entry_time': str(t.get('entry_time', '')),
            })

    all_trades.sort(key=lambda t: t['entry_time'])

    equity = capital
    peak   = equity
    max_dd = 0.0
    wins   = 0

    for t in all_trades:
        risk_amount = equity * (risk_pct / 100.0)
        outcome     = t['outcome']
        sl_pct      = max(t['sl_pct'], 0.01)

        if outcome == 'WIN':
            pnl = risk_amount * RR_RATIO
            wins += 1
        elif outcome == 'LOSS':
            pnl = -risk_amount
        else:  # TIMEOUT
            pnl = risk_amount * (t['pnl_pct'] / sl_pct)

        equity += pnl
        if equity > peak:
            peak = equity
        if peak > 0:
            dd = (peak - equity) / peak * 100.0
            if dd > max_dd:
                max_dd = dd

    n = len(all_trades)
    total_pnl_pct = (equity - capital) / capital * 100.0 if capital > 0 else 0.0

    return {
        'total_pnl_pct': total_pnl_pct,
        'final_equity':  equity,
        'max_dd':        max_dd,
        'n_trades':      n,
        'win_rate':      wins / n if n > 0 else 0.0,
    }


def compute_filtered_stats(trades: list, capital: float, risk_pct: float) -> dict:
    """Berechnet Einzel-Statistiken aus gefilterten Trades (shared-capital Modell)."""
    return simulate_portfolio(
        [{'market': '', 'timeframe': '', 'trades': trades}],
        capital, risk_pct
    )


def _calmar(metrics: dict) -> float:
    """Calmar Ratio: PnL% / MaxDD% (risiko-adjustierter Score, wie jaegerbot)."""
    if metrics['max_dd'] > 0:
        return metrics['total_pnl_pct'] / metrics['max_dd']
    return metrics['total_pnl_pct']


def optimize_portfolio(candidates: list, capital: float, risk_pct: float,
                        max_dd_limit: float) -> tuple:
    """
    Greedy-Algorithmus mit Calmar-Ratio-Score (wie jaegerbot).
    Kandidaten pro Iteration werden parallel via ThreadPoolExecutor bewertet.
    Constraint: max. 1 TF pro Coin (Bitget: 1 Position pro Symbol).
    """
    # Einzel-Stats für alle Kandidaten berechnen
    for r in candidates:
        r['filtered_stats'] = compute_filtered_stats(r['trades'], capital, risk_pct)

    # Besten Kandidaten pro Coin vorauswählen (höchster Calmar, MaxDD-konform)
    coin_best: dict = {}
    for r in candidates:
        st = r['filtered_stats']
        if st['total_pnl_pct'] <= 0 or st['max_dd'] > max_dd_limit:
            continue
        coin  = r['coin']
        score = _calmar(st)
        if coin not in coin_best or score > _calmar(coin_best[coin]['filtered_stats']):
            coin_best[coin] = r

    eligible = list(coin_best.values())
    if not eligible:
        return None, None

    eligible.sort(key=lambda r: _calmar(r['filtered_stats']), reverse=True)

    # Star-Spieler: Einzelstrategie mit höchstem Calmar
    best_team    = [eligible[0]]
    best_metrics = simulate_portfolio(best_team, capital, risk_pct)
    best_score   = _calmar(best_metrics)
    candidate_pool = eligible[1:]

    print(f"  1/3 Star-Spieler: {best_team[0]['market']} {best_team[0]['timeframe']} "
          f"(Calmar: {best_score:.2f})")
    print(f"  3/3 Suche beste Team-Kollegen "
          f"({len(candidate_pool)} Kandidaten, {N_WORKERS} Threads)...")

    while candidate_pool:
        best_addition     = None
        best_score_with   = best_score
        best_metrics_with = best_metrics

        # Snapshot des aktuellen Teams für parallele Ausführung
        current_team = list(best_team)

        def _eval(cand, _team=current_team):
            m = simulate_portfolio(_team + [cand], capital, risk_pct)
            if m['max_dd'] <= max_dd_limit:
                return cand, m, _calmar(m)
            return cand, None, -1.0

        with ThreadPoolExecutor(max_workers=N_WORKERS) as executor:
            futures = {executor.submit(_eval, c): c for c in candidate_pool}
            for future in as_completed(futures):
                cand, m, score = future.result()
                if m is not None and score > best_score_with:
                    best_score_with   = score
                    best_addition     = cand
                    best_metrics_with = m

        if best_addition:
            print(f"    + {best_addition['market']} {best_addition['timeframe']} "
                  f"(Calmar: {best_score_with:.2f})")
            best_team.append(best_addition)
            best_score    = best_score_with
            best_metrics  = best_metrics_with
            candidate_pool.remove(best_addition)
        else:
            print(f"  {Y}Keine weitere Verbesserung durch mehr Coins. Optimierung beendet.{NC}")
            break

    return best_metrics, best_team


def print_result(selected: list, pm: dict, capital: float, risk_pct: float,
                 max_dd_limit: float):
    w = 72
    print(f"\n{'=' * w}")
    print(f"{B}  dnabot — Automatische Portfolio-Optimierung{NC}")
    print(f"  Ziel: Maximaler Profit bei maximal {max_dd_limit:.1f}% Drawdown.")
    print(f"{'=' * w}")

    if not selected:
        print(f"\n{R}  Kein Portfolio gefunden, das die Bedingungen erfüllt.{NC}")
        print(f"  → Erhöhe das Max-Drawdown-Limit oder führe zuerst Mode 1 aus.\n")
        return

    print(f"\n  {G}Optimales Portfolio — {len(selected)} Coin(s){NC}")
    print(f"  Kapital: {capital:.0f} USDT | Risiko/Trade: {risk_pct}% (gemeinsamer Pool)")
    print(f"\n  {'Markt':<24} {'TF':<6} {'Trades':>7} {'WR':>7} {'PnL%':>9} {'MaxDD':>8}")
    print(f"  {'-' * (w - 2)}")

    for pr in sorted(selected, key=lambda x: x['filtered_stats']['total_pnl_pct'], reverse=True):
        st  = pr['filtered_stats']
        pnl_col = G if st['total_pnl_pct'] > 0 else R
        wr_col  = G if st['win_rate'] >= 0.50 else (Y if st['win_rate'] >= 0.43 else R)
        sign = '+' if st['total_pnl_pct'] >= 0 else ''
        print(
            f"  {pr['market']:<24} {pr['timeframe']:<6} {st['n_trades']:>7} "
            f"{wr_col}{st['win_rate']:>6.1%}{NC} "
            f"{pnl_col}{sign}{st['total_pnl_pct']:>7.1f}%{NC} "
            f"{st['max_dd']:>7.1f}%"
        )

    pnl_col = G if pm['total_pnl_pct'] > 0 else R
    calmar  = _calmar(pm)
    print(f"\n  {'─' * (w - 2)}")
    print(f"  {B}Portfolio gesamt (gemeinsamer Kapital-Pool, alle Trades kompoundiert):{NC}")
    print(f"  Trades total:  {pm['n_trades']}")
    print(f"  Win-Rate:      {pm['win_rate']:.1%}")
    print(f"  PnL:           {pnl_col}{'+' if pm['total_pnl_pct'] >= 0 else ''}{pm['total_pnl_pct']:.1f}%{NC}")
    print(f"  Final Equity:  {pm['final_equity']:.2f} USDT")
    print(f"  Max Drawdown:  {pm['max_dd']:.1f}%")
    print(f"  Calmar Score:  {G}{calmar:.2f}{NC}  (PnL% / MaxDD% — höher = besser)")
    print(f"{'=' * w}\n")


def generate_portfolio_equity_chart(selected: list, pm: dict,
                                     start_date: str, end_date: str,
                                     capital: float, risk_pct: float):
    """
    Erstellt einen kombinierten Portfolio-Equity-Chart im gleichen Stil wie Option 5.
    make_subplots(secondary_y=True):
      - Primäre Y-Achse (links):  Einzel-Equity pro Pair (dünne Linien)
      - Sekundäre Y-Achse (rechts): Kombinierte Portfolio-Equity (blaue Hauptlinie)
      - WIN/LOSS/TIMEOUT-Marker auf der Portfolio-Equity (gleiche Symbole wie Option 5)
    """
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        print(f"{R}  plotly nicht installiert — Chart übersprungen.{NC}")
        return

    # Telegram-Credentials
    secret_path = os.path.join(PROJECT_ROOT, 'secret.json')
    bot_token, chat_id = '', ''
    try:
        with open(secret_path) as f:
            secrets = json.load(f)
        accounts  = secrets.get('dnabot', [])
        tg        = accounts[0] if accounts else {}
        bot_token = tg.get('telegram_bot_token', '') or secrets.get('telegram', {}).get('bot_token', '')
        chat_id   = tg.get('telegram_chat_id', '')   or secrets.get('telegram', {}).get('chat_id', '')
    except Exception:
        pass

    # ── Alle Trades zusammenführen & chronologisch sortieren ────────────────
    all_trades = []
    for pr in selected:
        for t in pr['trades']:
            all_trades.append({
                'market':     pr['market'],
                'timeframe':  pr['timeframe'],
                'outcome':    t.get('outcome', 'LOSS'),
                'pnl_pct':    t.get('pnl_pct', 0.0),
                'sl_pct':     t.get('sl_pct', 1.0),
                'entry_time': str(t.get('entry_time', '')),
            })
    all_trades.sort(key=lambda t: t['entry_time'])

    if not all_trades:
        print(f"  {Y}Keine Trades vorhanden — Chart übersprungen.{NC}")
        return

    # ── Kombinierte Portfolio-Equity ────────────────────────────────────────
    equity    = capital
    peak      = equity
    eq_times  = [all_trades[0]['entry_time']]
    eq_vals   = [equity]
    wins      = 0

    for t in all_trades:
        risk_amount = equity * (risk_pct / 100.0)
        outcome     = t['outcome']
        sl_pct      = max(t['sl_pct'], 0.01)
        if outcome == 'WIN':
            equity += risk_amount * RR_RATIO
            wins   += 1
        elif outcome == 'LOSS':
            equity -= risk_amount
        else:
            equity += risk_amount * (t['pnl_pct'] / sl_pct)
        if equity > peak:
            peak = equity
        eq_times.append(t['entry_time'])
        eq_vals.append(round(equity, 2))

    n       = len(all_trades)
    wr      = wins / n if n > 0 else 0.0
    pnl_pct = (equity - capital) / capital * 100.0 if capital > 0 else 0.0
    max_dd  = pm['max_dd']
    sign    = '+' if pnl_pct >= 0 else ''

    # ── Einzel-Equity pro Pair (für primäre Y-Achse) ────────────────────────
    PAIR_COLORS = [
        '#f59e0b', '#8b5cf6', '#ec4899', '#14b8a6',
        '#f97316', '#84cc16', '#06b6d4', '#a78bfa',
    ]
    pair_equity_traces = []
    for idx, pr in enumerate(selected):
        pair_trades = sorted(
            [t for t in pr['trades']],
            key=lambda t: str(t.get('entry_time', ''))
        )
        peq    = capital
        ptimes = [str(pair_trades[0].get('entry_time', ''))] if pair_trades else []
        pvals  = [peq]
        for t in pair_trades:
            ra  = peq * (risk_pct / 100.0)
            out = t.get('outcome', 'LOSS')
            slp = max(t.get('sl_pct', 1.0), 0.01)
            if out == 'WIN':
                peq += ra * RR_RATIO
            elif out == 'LOSS':
                peq -= ra
            else:
                peq += ra * (t.get('pnl_pct', 0.0) / slp)
            ptimes.append(str(t.get('entry_time', '')))
            pvals.append(round(peq, 2))
        label = f"{pr['market'].split('/')[0]}/{pr['timeframe']}"
        pair_equity_traces.append((ptimes, pvals, label, PAIR_COLORS[idx % len(PAIR_COLORS)]))

    # ── Figure aufbauen (identischer Aufbau wie create_chart in Option 5) ───
    date_range = f"{start_date or '...'} → {end_date or 'heute'}"
    pairs_str  = ', '.join(f"{p['market'].split('/')[0]}/{p['timeframe']}" for p in selected)
    title = (
        f"dnabot Portfolio — {len(selected)} Coins ({pairs_str}) | "
        f"Trades: {n} | WR: {wr:.1%} | "
        f"PnL: {sign}{pnl_pct:.1f}% | "
        f"Final Equity: {equity:.2f} USDT | MaxDD: {max_dd:.1f}%"
    )

    fig = make_subplots(specs=[[{"secondary_y": True}]])

    # Einzel-Equity-Linien (primäre Y-Achse, dünn)
    for ptimes, pvals, label, color in pair_equity_traces:
        fig.add_trace(go.Scatter(
            x=ptimes, y=pvals,
            mode='lines',
            name=label,
            line=dict(color=color, width=1),
            opacity=0.55,
        ), secondary_y=False)

    # Startkapital-Referenzlinie
    fig.add_hline(
        y=capital,
        line=dict(color='rgba(100,100,100,0.35)', width=1, dash='dash'),
        annotation_text=f'Start {capital:.0f} USDT',
        annotation_position='top left',
    )

    # ── WIN / LOSS / TIMEOUT Marker auf Portfolio-Equity (sekundäre Y) ──────
    entry_long_x,  entry_long_y,  entry_long_txt  = [], [], []
    entry_short_x, entry_short_y, entry_short_txt = [], [], []
    exit_win_x,    exit_win_y    = [], []
    exit_loss_x,   exit_loss_y   = [], []
    exit_to_x,     exit_to_y     = [], []

    for i, t in enumerate(all_trades):
        eq_val = eq_vals[i + 1]
        tip    = f"{t['market']} {t['timeframe']}<br>Equity: {eq_val:.2f} USDT"
        # Alle Trades als Entry Long aufführen (kein direction im portfolio-trade)
        entry_long_x.append(t['entry_time'])
        entry_long_y.append(eq_val)
        entry_long_txt.append(tip)
        if t['outcome'] == 'WIN':
            exit_win_x.append(t['entry_time']);  exit_win_y.append(eq_val)
        elif t['outcome'] == 'LOSS':
            exit_loss_x.append(t['entry_time']); exit_loss_y.append(eq_val)
        else:
            exit_to_x.append(t['entry_time']);   exit_to_y.append(eq_val)

    # Portfolio-Equity-Kurve (sekundäre Y-Achse, blau wie Option 5)
    fig.add_trace(go.Scatter(
        x=eq_times, y=eq_vals,
        mode='lines',
        name='Portfolio Equity',
        line=dict(color='#2563eb', width=2),
        opacity=0.75,
    ), secondary_y=True)

    # Entry-Marker (▲ grün / ▼ orange — gleiche Symbole wie Option 5)
    if entry_long_x:
        fig.add_trace(go.Scatter(
            x=entry_long_x, y=entry_long_y, mode='markers',
            marker=dict(color='#16a34a', symbol='triangle-up', size=14,
                        line=dict(width=1, color='#0f5132')),
            name='Entry ▲', text=entry_long_txt,
            hovertemplate='%{text}<extra>Entry</extra>',
        ), secondary_y=True)

    # Exit WIN ● cyan
    if exit_win_x:
        fig.add_trace(go.Scatter(
            x=exit_win_x, y=exit_win_y, mode='markers',
            marker=dict(color='#22d3ee', symbol='circle', size=11,
                        line=dict(width=1, color='#0e7490')),
            name='Exit TP ✓',
        ), secondary_y=True)

    # Exit LOSS ✗ rot
    if exit_loss_x:
        fig.add_trace(go.Scatter(
            x=exit_loss_x, y=exit_loss_y, mode='markers',
            marker=dict(color='#ef4444', symbol='x', size=11,
                        line=dict(width=2, color='#7f1d1d')),
            name='Exit SL ✗',
        ), secondary_y=True)

    # Exit TIMEOUT ■ grau
    if exit_to_x:
        fig.add_trace(go.Scatter(
            x=exit_to_x, y=exit_to_y, mode='markers',
            marker=dict(color='#9ca3af', symbol='square', size=9),
            name='Exit Timeout',
        ), secondary_y=True)

    # ── Layout — identisch zu create_chart() in Option 5 ────────────────────
    fig.update_layout(
        title=dict(text=title, font=dict(size=13), x=0.5, xanchor='center'),
        height=750,
        hovermode='x unified',
        template='plotly_white',
        dragmode='zoom',
        xaxis=dict(rangeslider=dict(visible=True), fixedrange=False),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
    )
    fig.update_yaxes(title_text='Einzel-Equity (USDT)', secondary_y=False, fixedrange=False)
    fig.update_yaxes(title_text='Portfolio-Equity (USDT)', secondary_y=True,  fixedrange=False)

    output_file = '/tmp/dnabot_portfolio_equity.html'
    fig.write_html(output_file)
    print(f"\n  {G}✓ Portfolio-Chart erstellt: {output_file}{NC}")

    if bot_token and chat_id:
        sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))
        from dnabot.utils.telegram import send_document
        caption = (
            f"dnabot Portfolio-Equity\n"
            f"{date_range} | {len(selected)} Coins | "
            f"PnL: {sign}{pnl_pct:.1f}% | Equity: {equity:.2f} USDT | MaxDD: {max_dd:.1f}%"
        )
        send_document(bot_token, chat_id, output_file, caption=caption)
        print(f"  {G}✓ Via Telegram gesendet.{NC}")
    else:
        print(f"  {Y}Telegram nicht konfiguriert — Chart nur lokal gespeichert.{NC}")


def generate_trades_excel(selected: list, pm: dict, capital: float, risk_pct: float,
                          leverage: int = 1):
    """Erstellt eine Excel-Tabelle mit allen Einzeltrades des optimalen Portfolios."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"  {Y}openpyxl nicht installiert — Excel übersprungen. (pip install openpyxl){NC}")
        return None

    # Alle Trades zusammenführen und chronologisch sortieren
    all_trades = []
    for pr in selected:
        for t in pr['trades']:
            all_trades.append({
                'market':     pr['market'],
                'timeframe':  pr['timeframe'],
                'coin':       pr['market'].split('/')[0],
                'direction':  t.get('direction', '?'),
                'outcome':    t.get('outcome', 'LOSS'),
                'pnl_pct':    t.get('pnl_pct', 0.0),
                'sl_pct':     t.get('sl_pct', 1.0),
                'entry_time': str(t.get('entry_time', '')),
                'exit_time':  str(t.get('exit_time', '')),
            })
    all_trades.sort(key=lambda t: t['entry_time'])

    # Equity-Verlauf berechnen
    equity = capital
    rows = []
    for i, t in enumerate(all_trades):
        equity_before = equity
        risk_amount   = equity_before * (risk_pct / 100.0)
        sl_pct        = max(t['sl_pct'], 0.01)
        outcome       = t['outcome']
        if outcome == 'WIN':
            pnl = risk_amount * RR_RATIO
        elif outcome == 'LOSS':
            pnl = -risk_amount
        else:
            pnl = risk_amount * (t['pnl_pct'] / sl_pct)
        equity += pnl

        # Marge = Positionsgröße / Leverage, gedeckelt auf verfügbares Kapital
        raw_position = risk_amount / max(t.get('sl_pct', 1.0) / 100.0, 0.0001)
        max_position = equity_before * max(leverage, 1)
        margin = min(raw_position, max_position) / max(leverage, 1)

        ergebnis = 'TP erreicht' if outcome == 'WIN' else ('SL erreicht' if outcome == 'LOSS' else 'Timeout')
        rows.append({
            'Nr':                    i + 1,
            'Datum':                 t['entry_time'][:16].replace('T', ' '),
            'Coin':                  t['coin'],
            'Timeframe':             t['timeframe'],
            'Richtung':              t['direction'],
            'Ergebnis':              ergebnis,
            'Reale Bewegung (%)':    round(t.get('pnl_pct', 0.0), 4),
            'Riskiert (USDT)':       round(risk_amount, 4),
            'Marge (USDT)':          round(margin, 4),
            'PnL (USDT)':            round(pnl, 4),
            'Gesamtkapital':         round(equity, 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'

    # Farben
    header_fill  = PatternFill('solid', fgColor='1E3A5F')
    win_fill     = PatternFill('solid', fgColor='D6F4DC')
    loss_fill    = PatternFill('solid', fgColor='FAD7D7')
    timeout_fill = PatternFill('solid', fgColor='FFF3CC')
    alt_fill     = PatternFill('solid', fgColor='F2F2F2')
    thin_border  = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    headers = list(rows[0].keys()) if rows else []
    col_widths = {
        'Nr': 6, 'Datum': 18, 'Coin': 10, 'Timeframe': 12,
        'Richtung': 10, 'Ergebnis': 14, 'Reale Bewegung (%)': 20,
        'Riskiert (USDT)': 16, 'Marge (USDT)': 14, 'PnL (USDT)': 14, 'Gesamtkapital': 16,
    }

    # Header
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 14)
    ws.row_dimensions[1].height = 22

    # Datenzeilen
    for r_idx, row in enumerate(rows, 2):
        outcome_val = row['Ergebnis']
        if outcome_val == 'TP erreicht':
            fill = win_fill
        elif outcome_val == 'SL erreicht':
            fill = loss_fill
        else:
            fill = timeout_fill if r_idx % 2 == 0 else alt_fill

        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=col, value=row[key])
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Reale Bewegung (%)', 'Riskiert (USDT)', 'Marge (USDT)', 'PnL (USDT)', 'Gesamtkapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[r_idx].height = 18

    # Zusammenfassung unten
    summary_row = len(rows) + 3
    ws.cell(row=summary_row, column=1, value='Zusammenfassung').font = Font(bold=True, size=11)
    for label, value in [
        ('Trades gesamt', pm['n_trades']),
        ('Win-Rate', f"{pm['win_rate']:.1%}"),
        ('PnL', f"+{pm['total_pnl_pct']:.1f}%"),
        ('Final Equity', f"{pm['final_equity']:.2f} USDT"),
        ('Max Drawdown', f"{pm['max_dd']:.1f}%"),
        ('Risiko/Trade', f"{risk_pct}%"),
    ]:
        ws.cell(row=summary_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=value)
        summary_row += 1

    output_file = '/tmp/dnabot_trades.xlsx'
    wb.save(output_file)
    print(f"  {G}✓ Excel-Tabelle erstellt: {output_file}{NC}")
    return output_file


def write_to_settings(selected: list, risk_pct: float = None):
    try:
        with open(SETTINGS_PATH) as f:
            settings = json.load(f)
    except Exception as e:
        print(f"{R}Fehler beim Lesen von settings.json: {e}{NC}")
        return False

    new_strategies = [
        {"symbol": pr['market'], "timeframe": pr['timeframe'], "active": True}
        for pr in selected
    ]
    settings.setdefault('live_trading_settings', {})['active_strategies'] = new_strategies

    if risk_pct is not None:
        settings.setdefault('risk_settings', {})['risk_per_entry_pct'] = risk_pct

    try:
        with open(SETTINGS_PATH, 'w') as f:
            json.dump(settings, f, indent=2)
        risk_info = f" | Risiko/Trade: {risk_pct}%" if risk_pct is not None else ""
        print(f"\n{G}✓ settings.json aktualisiert — {len(new_strategies)} Strategie(n) eingetragen{risk_info}.{NC}\n")
        return True
    except Exception as e:
        print(f"{R}Fehler beim Schreiben von settings.json: {e}{NC}")
        return False


def main():
    parser = argparse.ArgumentParser(description="dnabot Portfolio Optimizer")
    parser.add_argument('--capital',    type=float, default=1000.0)
    parser.add_argument('--risk',       type=float, default=1.0)
    parser.add_argument('--max-dd',     type=float, default=30.0)
    parser.add_argument('--start-date', type=str,   default=None)
    parser.add_argument('--end-date',   type=str,   default=None)
    parser.add_argument('--auto-write', action='store_true')
    args = parser.parse_args()

    date_range = ""
    if args.start_date or args.end_date:
        date_range = f" | {args.start_date or '...'} → {args.end_date or 'heute'}"

    print(f"\n{'─' * 72}")
    print(f"{B}  dnabot Automatische Portfolio-Optimierung{NC}")
    print(f"  Ziel: Maximaler Profit bei maximal {args.max_dd:.1f}% Drawdown.{date_range}")
    print(f"  Modell: Gemeinsamer Kapital-Pool — alle Trades kompoundieren zusammen")
    print(f"  Constraint: max. 1 Timeframe pro Coin (Bitget-Regel)")
    print(f"{'─' * 72}\n")

    print("  Lade Backtest-Ergebnisse ...", end='', flush=True)
    all_results = load_all_results(args.start_date, args.end_date)
    if not all_results:
        print(f"\n{R}  Keine Backtest-Ergebnisse gefunden.{NC}")
        print("  Zuerst Mode 1 (Einzel-Backtest) ausführen!\n")
        sys.exit(1)

    with_trades = [r for r in all_results if len(r['trades']) > 0]
    coins = sorted(set(r['coin'] for r in with_trades))
    print(f" {len(all_results)} Dateien, {len(with_trades)} mit Trades, {len(coins)} Coins.")

    # Risiko 1%–5% (Schritte 0.5%) ausprobieren, bestes Final Equity unter MaxDD-Limit wählen
    risk_levels = [r / 10 for r in range(10, 55, 5)]  # 1.0, 1.5, 2.0, ... 5.0
    print(f"\n  Suche optimales Risiko ({risk_levels[0]}%–{risk_levels[-1]}%, MaxDD ≤ {args.max_dd:.0f}%)...\n")

    best_metrics   = None
    best_combo     = None
    best_risk      = args.risk
    best_equity    = 0.0
    best_calmar    = -999.0

    for risk_pct in risk_levels:
        m, combo = optimize_portfolio(with_trades, args.capital, risk_pct, args.max_dd)
        if combo and m and m['max_dd'] <= args.max_dd:
            score = _calmar(m)
            if score > best_calmar:
                best_metrics = m
                best_combo   = combo
                best_risk    = risk_pct
                best_equity  = m['final_equity']
                best_calmar  = score

    if not best_combo:
        best_metrics = {'total_pnl_pct': 0, 'final_equity': args.capital,
                        'max_dd': 0, 'n_trades': 0, 'win_rate': 0}
        best_combo = []

    print(f"\n  {G}Bestes Risiko: {best_risk}% → Calmar: {best_calmar:.2f} | Final Equity: {best_equity:.2f} USDT{NC}\n")
    print_result(best_combo, best_metrics, args.capital, best_risk, args.max_dd)

    if not best_combo:
        sys.exit(0)

    # Aktuelles Portfolio aus settings.json simulieren zum Vergleich
    current_equity = 0.0
    current_capital = None
    leverage = 1
    try:
        with open(SETTINGS_PATH) as f:
            current_settings = json.load(f)
        current_capital = current_settings.get('optimization_settings', {}).get('start_capital')
        leverage = int(current_settings.get('risk_settings', {}).get('leverage', 1))
        current_strategies = current_settings.get('live_trading_settings', {}).get('active_strategies', [])
        if current_strategies:
            current_pairs = []
            for strat in current_strategies:
                sym = strat.get('symbol', '')
                tf  = strat.get('timeframe', '')
                match = next((r for r in with_trades if r['market'] == sym and r['timeframe'] == tf), None)
                if match:
                    current_pairs.append(match)
            if current_pairs:
                for r in current_pairs:
                    r['filtered_stats'] = compute_filtered_stats(r['trades'], args.capital, best_risk)
                sim = simulate_portfolio(current_pairs, args.capital, best_risk)
                current_equity = sim['final_equity']
    except Exception:
        pass

    new_pairs_str = ', '.join(
        f"{p['market'].split('/')[0]} ({p['timeframe']})" for p in best_combo
    )

    # Kapitaländerung = immer überschreiben
    capital_changed = current_capital is not None and float(current_capital) != float(args.capital)

    # settings.json Entscheidung
    settings_updated = False
    if current_equity > 0 and not capital_changed:
        print(f"  Aktuelles Portfolio @ {best_risk}%: {current_equity:.2f} USDT")
        if best_equity <= current_equity:
            print(f"  {Y}Neues Ergebnis ({best_equity:.2f} USDT) ist nicht besser → settings.json bleibt unverändert.{NC}\n")
        else:
            print(f"  {G}Verbesserung: {current_equity:.2f} → {best_equity:.2f} USDT → überschreibe settings.json{NC}\n")
            if args.auto_write:
                write_to_settings(best_combo, best_risk)
                settings_updated = True
            else:
                try:
                    ans = input("  Sollen die optimalen Ergebnisse automatisch in settings.json eingetragen werden? (j/n): ").strip().lower()
                except (EOFError, KeyboardInterrupt):
                    ans = 'n'
                if ans in ('j', 'ja', 'y', 'yes'):
                    write_to_settings(best_combo, best_risk)
                    settings_updated = True
                else:
                    print(f"\n{Y}  settings.json wurde NICHT geändert.{NC}\n")
    elif capital_changed:
        print(f"  {Y}Kapital geändert ({current_capital} → {args.capital} USDT) → überschreibe settings.json.{NC}\n")
        write_to_settings(best_combo, best_risk)
        settings_updated = True
    else:
        write_to_settings(best_combo, best_risk)
        settings_updated = True

    # Telegram: Ergebnis der settings-Entscheidung
    if args.auto_write:
        if settings_updated:
            _send_telegram(
                f"dnabot Auto-Optimizer — Portfolio aktualisiert\n"
                f"Equity: {current_equity:.2f} → {best_equity:.2f} USDT (+{((best_equity/max(current_equity,0.01))-1)*100:.1f}%)\n"
                f"Risiko: {best_risk}% | MaxDD: {best_metrics['max_dd']:.1f}% | WR: {best_metrics['win_rate']:.1%}\n"
                f"Neue Coins:\n{new_pairs_str}"
            )
        else:
            _send_telegram(
                f"dnabot Auto-Optimizer — Keine Änderung\n"
                f"Neues Portfolio ({best_equity:.2f} USDT) ist nicht besser als aktuelles ({current_equity:.2f} USDT).\n"
                f"Risiko: {best_risk}% | MaxDD: {best_metrics['max_dd']:.1f}%\n"
                f"settings.json bleibt unverändert."
            )

    # Charts + Excel: bei --auto-write immer, sonst interaktiv fragen
    if args.auto_write:
        generate_portfolio_equity_chart(
            best_combo, best_metrics, args.start_date, args.end_date, args.capital, best_risk
        )
        excel_file = generate_trades_excel(best_combo, best_metrics, args.capital, best_risk, leverage)
        if excel_file:
            bot_token, chat_id = _get_telegram_credentials()
            if bot_token and chat_id:
                sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))
                from dnabot.utils.telegram import send_document
                send_document(
                    bot_token, chat_id, excel_file,
                    caption=f"dnabot Trades-Tabelle | {len(best_combo)} Coins | "
                            f"Risiko: {best_risk}% | {best_metrics['n_trades']} Trades | "
                            f"WR: {best_metrics['win_rate']:.1%} | Final: {best_metrics['final_equity']:.2f} USDT"
                )
                print(f"  {G}✓ Excel via Telegram gesendet.{NC}")
    else:
        try:
            chart_ans = input("  Interaktive Charts für diese Zusammenstellung erstellen & via Telegram senden? (j/n): ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            chart_ans = 'n'
        if chart_ans in ('j', 'ja', 'y', 'yes'):
            generate_portfolio_equity_chart(
                best_combo, best_metrics, args.start_date, args.end_date, args.capital, best_risk
            )
            excel_file = generate_trades_excel(best_combo, best_metrics, args.capital, best_risk, leverage)
            if excel_file:
                bot_token, chat_id = _get_telegram_credentials()
                if bot_token and chat_id:
                    sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))
                    from dnabot.utils.telegram import send_document
                    send_document(
                        bot_token, chat_id, excel_file,
                        caption=f"dnabot Trades-Tabelle | {len(best_combo)} Coins | "
                                f"Risiko: {best_risk}% | {best_metrics['n_trades']} Trades | "
                                f"WR: {best_metrics['win_rate']:.1%} | Final: {best_metrics['final_equity']:.2f} USDT"
                    )
                    print(f"  {G}✓ Excel via Telegram gesendet.{NC}")


if __name__ == '__main__':
    main()
